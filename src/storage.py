from __future__ import annotations

import csv
import hashlib
import json
import sqlite3
import shutil
from pathlib import Path
from typing import Any

from .schema import flatten_document_row, flatten_raw_fields


SUMMARY_HEADERS = [
    "task_id",
    "file_hash",
    "document_type",
    "document_no",
    "document_date",
    "sender",
    "receiver",
    "total_quantity",
    "item_count",
    "first_item_name",
    "first_item_code",
    "confidence",
    "source_image",
    "processed_at",
    "items_json",
    "raw_text",
    "raw_model_json",
]

ITEM_HEADERS = [
    "task_id",
    "file_hash",
    "table_name",
    "row_index",
    "row_json",
    "source_image",
    "processed_at",
]

RAW_FIELD_HEADERS = [
    "task_id",
    "file_hash",
    "field_path",
    "value",
    "value_json",
    "source_image",
    "processed_at",
]

RAW_JSON_HEADERS = [
    "task_id",
    "file_hash",
    "raw_model_json",
    "source_image",
    "processed_at",
]

MODEL_SUMMARY_HEADERS = [
    "task_id",
    "file_hash",
    "source_image",
    "processed_at",
    "document_category",
    "common_fields_json",
    "raw_fields_json",
    "extra_fields_json",
]

MODEL_TABLE_HEADERS = [
    "task_id",
    "file_hash",
    "source_image",
    "processed_at",
    "table_name",
    "row_index",
    "row_json",
]

MODEL_RAW_FIELD_HEADERS = [
    "task_id",
    "file_hash",
    "source_image",
    "processed_at",
    "field_path",
    "value",
    "value_json",
]

MODEL_RAW_JSON_HEADERS = [
    "task_id",
    "file_hash",
    "source_image",
    "processed_at",
    "model_json",
]


class StorageManager:
    def __init__(self, root_dir: Path, config: dict[str, Any]) -> None:
        self.root_dir = root_dir
        self.config = config
        app_config = config.get("app", {})

        self.input_dir = (root_dir / app_config.get("input_dir", "data/input")).resolve()
        self.output_dir = (root_dir / app_config.get("output_dir", "data/output")).resolve()
        self.model_json_dir = (root_dir / app_config.get("model_json_dir", "data/output/model_json")).resolve()
        self.debug_dir = (root_dir / app_config.get("debug_dir", "data/output/debug")).resolve()
        self.archive_dir = (root_dir / app_config.get("archive_dir", "data/archive")).resolve()
        self.logs_dir = (root_dir / app_config.get("logs_dir", "data/logs")).resolve()

        self.json_dir = self.output_dir / "json"
        self.tables_dir = self.output_dir / "tables"
        self.success_dir = self.archive_dir / "success"
        self.error_dir = self.archive_dir / "error"
        self.task_logs_dir = self.logs_dir / "tasks"

        self.summary_csv = (root_dir / app_config.get("summary_csv", "data/output/tables/summary.csv")).resolve()
        self.summary_xlsx = (root_dir / app_config.get("summary_xlsx", "data/output/tables/summary.xlsx")).resolve()
        self.model_results_xlsx = (root_dir / app_config.get("model_results_xlsx", "data/output/tables/model_results.xlsx")).resolve()
        self.state_db = (root_dir / app_config.get("state_db", "data/logs/state.db")).resolve()

        self.ensure_directories()
        self._init_db()

    def ensure_directories(self) -> None:
        for path in (
            self.input_dir,
            self.output_dir,
            self.model_json_dir,
            self.debug_dir,
            self.json_dir,
            self.tables_dir,
            self.archive_dir,
            self.success_dir,
            self.error_dir,
            self.logs_dir,
            self.task_logs_dir,
        ):
            path.mkdir(parents=True, exist_ok=True)

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.state_db)
        conn.row_factory = sqlite3.Row
        return conn

    def _init_db(self) -> None:
        with self._connect() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS processed_documents (
                    file_hash TEXT PRIMARY KEY,
                    task_id TEXT NOT NULL,
                    source_path TEXT NOT NULL,
                    status TEXT NOT NULL,
                    output_json_path TEXT,
                    archived_path TEXT,
                    error_message TEXT,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS task_logs (
                    task_id TEXT PRIMARY KEY,
                    file_hash TEXT NOT NULL,
                    source_path TEXT NOT NULL,
                    status TEXT NOT NULL,
                    payload_json TEXT,
                    error_message TEXT,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            conn.commit()

    @staticmethod
    def sha256_file(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    def is_processed_hash(self, file_hash: str) -> bool:
        with self._connect() as conn:
            row = conn.execute(
                "SELECT 1 FROM processed_documents WHERE file_hash = ?",
                (file_hash,),
            ).fetchone()
        return row is not None

    def get_processed_record(self, file_hash: str) -> dict[str, Any] | None:
        with self._connect() as conn:
            row = conn.execute(
                "SELECT * FROM processed_documents WHERE file_hash = ?",
                (file_hash,),
            ).fetchone()
        return dict(row) if row else None

    def store_task_log(
        self,
        *,
        task_id: str,
        file_hash: str,
        source_path: str,
        status: str,
        payload: dict[str, Any] | None,
        error_message: str | None,
    ) -> None:
        payload_json = json.dumps(payload, ensure_ascii=False, indent=2) if payload is not None else None
        with self._connect() as conn:
            conn.execute(
                """
                INSERT INTO task_logs(task_id, file_hash, source_path, status, payload_json, error_message, updated_at)
                VALUES(?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(task_id) DO UPDATE SET
                    file_hash=excluded.file_hash,
                    source_path=excluded.source_path,
                    status=excluded.status,
                    payload_json=excluded.payload_json,
                    error_message=excluded.error_message,
                    updated_at=CURRENT_TIMESTAMP
                """,
                (task_id, file_hash, source_path, status, payload_json, error_message),
            )
            conn.commit()

    def store_processed_record(
        self,
        *,
        task_id: str,
        file_hash: str,
        source_path: str,
        status: str,
        output_json_path: str | None,
        archived_path: str | None,
        error_message: str | None,
    ) -> None:
        with self._connect() as conn:
            conn.execute(
                """
                INSERT INTO processed_documents(
                    file_hash, task_id, source_path, status, output_json_path, archived_path, error_message, updated_at
                )
                VALUES(?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(file_hash) DO UPDATE SET
                    task_id=excluded.task_id,
                    source_path=excluded.source_path,
                    status=excluded.status,
                    output_json_path=excluded.output_json_path,
                    archived_path=excluded.archived_path,
                    error_message=excluded.error_message,
                    updated_at=CURRENT_TIMESTAMP
                """,
                (
                    file_hash,
                    task_id,
                    source_path,
                    status,
                    output_json_path,
                    archived_path,
                    error_message,
                ),
            )
            conn.commit()

    def write_json_document(self, task_id: str, document: dict[str, Any]) -> Path:
        path = self.json_dir / f"{task_id}.json"
        path.write_text(json.dumps(document, ensure_ascii=False, indent=2), encoding="utf-8")
        return path

    def write_model_json_document(self, task_id: str, document: dict[str, Any]) -> Path:
        path = self.model_json_dir / f"{task_id}.json"
        path.write_text(json.dumps(document, ensure_ascii=False, indent=2), encoding="utf-8")
        return path

    def write_debug_document(self, task_id: str, document: dict[str, Any]) -> Path:
        path = self.debug_dir / f"{task_id}.json"
        path.write_text(json.dumps(document, ensure_ascii=False, indent=2), encoding="utf-8")
        return path

    def append_model_result(
        self,
        *,
        task_id: str,
        file_hash: str,
        source_image: str,
        processed_at: str,
        model_json: dict[str, Any],
    ) -> Path | None:
        try:
            from openpyxl import Workbook, load_workbook
        except Exception:
            return None

        self.tables_dir.mkdir(parents=True, exist_ok=True)
        if self.model_results_xlsx.exists():
            workbook = load_workbook(self.model_results_xlsx)
        else:
            workbook = Workbook()
            workbook.active.title = "model_summary"
            workbook["model_summary"].append(MODEL_SUMMARY_HEADERS)

        self._ensure_sheet(workbook, "model_summary", MODEL_SUMMARY_HEADERS)
        self._ensure_sheet(workbook, "model_tables", MODEL_TABLE_HEADERS)
        self._ensure_sheet(workbook, "model_raw_fields", MODEL_RAW_FIELD_HEADERS)
        self._ensure_sheet(workbook, "model_raw_json", MODEL_RAW_JSON_HEADERS)

        common_fields = model_json.get("common_fields") if isinstance(model_json.get("common_fields"), dict) else {}
        raw_fields = model_json.get("raw_fields") if isinstance(model_json.get("raw_fields"), dict) else {}
        extra_fields = model_json.get("extra_fields") if isinstance(model_json.get("extra_fields"), dict) else {}

        workbook["model_summary"].append(
            [
                task_id,
                file_hash,
                source_image,
                processed_at,
                model_json.get("document_category"),
                json.dumps(common_fields, ensure_ascii=False),
                json.dumps(raw_fields, ensure_ascii=False),
                json.dumps(extra_fields, ensure_ascii=False),
            ]
        )

        tables = model_json.get("tables") if isinstance(model_json.get("tables"), list) else []
        for table in tables:
            if not isinstance(table, dict):
                continue
            table_name = table.get("table_name") or table.get("name") or "table"
            rows = table.get("rows") if isinstance(table.get("rows"), list) else []
            for index, row in enumerate(rows, start=1):
                workbook["model_tables"].append(
                    [
                        task_id,
                        file_hash,
                        source_image,
                        processed_at,
                        table_name,
                        index,
                        json.dumps(row, ensure_ascii=False),
                    ]
                )

        for raw_row in flatten_raw_fields(model_json):
            workbook["model_raw_fields"].append(
                [
                    task_id,
                    file_hash,
                    source_image,
                    processed_at,
                    raw_row.get("field_path"),
                    raw_row.get("value"),
                    raw_row.get("value_json"),
                ]
            )

        workbook["model_raw_json"].append(
            [
                task_id,
                file_hash,
                source_image,
                processed_at,
                json.dumps(model_json, ensure_ascii=False),
            ]
        )

        workbook.save(self.model_results_xlsx)
        return self.model_results_xlsx

    @staticmethod
    def _ensure_sheet(workbook, name: str, headers: list[str]) -> None:
        if name in workbook.sheetnames:
            sheet = workbook[name]
            if sheet.max_row == 0:
                sheet.append(headers)
            return
        sheet = workbook.create_sheet(name)
        sheet.append(headers)

    def write_task_details(self, task_id: str, details: dict[str, Any]) -> Path:
        path = self.task_logs_dir / f"{task_id}.json"
        path.write_text(json.dumps(details, ensure_ascii=False, indent=2), encoding="utf-8")
        return path

    def append_summary_row(self, document: dict[str, Any]) -> tuple[Path, Path | None]:
        row = flatten_document_row(document)
        csv_path = self._append_csv_row(row)
        xlsx_path = self._append_xlsx_row(row)
        return csv_path, xlsx_path

    def _append_csv_row(self, row: dict[str, Any]) -> Path:
        self.tables_dir.mkdir(parents=True, exist_ok=True)
        target = self._summary_csv_target()
        file_exists = target.exists()
        with target.open("a", newline="", encoding="utf-8-sig") as handle:
            writer = csv.DictWriter(handle, fieldnames=SUMMARY_HEADERS)
            if not file_exists:
                writer.writeheader()
            writer.writerow({key: row.get(key) for key in SUMMARY_HEADERS})
        return target

    def _summary_csv_target(self) -> Path:
        if not self.summary_csv.exists():
            return self.summary_csv
        with self.summary_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            header = next(reader, [])
        if header == SUMMARY_HEADERS:
            return self.summary_csv
        return self.summary_csv.with_name(f"{self.summary_csv.stem}_v2{self.summary_csv.suffix}")

    def _append_xlsx_row(self, row: dict[str, Any]) -> Path | None:
        try:
            from openpyxl import Workbook, load_workbook
        except Exception:
            return None

        self.tables_dir.mkdir(parents=True, exist_ok=True)
        target = self._summary_xlsx_target()
        if target.exists():
            workbook = load_workbook(target)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "summary"
            sheet.append(SUMMARY_HEADERS)

        if sheet.max_row == 1 and all(sheet.cell(1, idx + 1).value == header for idx, header in enumerate(SUMMARY_HEADERS)):
            pass
        elif sheet.max_row == 0:
            sheet.append(SUMMARY_HEADERS)

        sheet.append([row.get(key) for key in SUMMARY_HEADERS])
        self._append_table_rows(workbook, row)
        self._append_raw_field_rows(workbook, row)
        self._append_raw_json_row(workbook, row)
        workbook.save(target)
        return target

    def _append_table_rows(self, workbook, summary_row: dict[str, Any]) -> None:
        if "tables" in workbook.sheetnames:
            sheet = workbook["tables"]
        else:
            sheet = workbook.create_sheet("tables")
            sheet.append(ITEM_HEADERS)

        try:
            raw_payload = json.loads(summary_row.get("raw_model_json") or "{}")
        except json.JSONDecodeError:
            raw_payload = {}

        tables = raw_payload.get("tables")
        if not isinstance(tables, list):
            tables = []
        if not tables:
            fallback_rows = raw_payload.get("line_items") or raw_payload.get("items") or []
            if fallback_rows:
                tables = [{"table_name": "line_items", "rows": fallback_rows}]

        for table in tables:
            if not isinstance(table, dict):
                continue
            table_name = table.get("table_name") or table.get("name") or "table"
            rows = table.get("rows")
            if not isinstance(rows, list):
                continue
            for index, row in enumerate(rows, start=1):
                sheet.append(
                    [
                        summary_row.get("task_id"),
                        summary_row.get("file_hash"),
                        table_name,
                        index,
                        json.dumps(row, ensure_ascii=False),
                        summary_row.get("source_image"),
                        summary_row.get("processed_at"),
                    ]
                )

    def _append_raw_field_rows(self, workbook, summary_row: dict[str, Any]) -> None:
        if "raw_fields" in workbook.sheetnames:
            sheet = workbook["raw_fields"]
        else:
            sheet = workbook.create_sheet("raw_fields")
            sheet.append(RAW_FIELD_HEADERS)

        try:
            raw_payload = json.loads(summary_row.get("raw_model_json") or "{}")
        except json.JSONDecodeError:
            raw_payload = {"raw_model_json": summary_row.get("raw_model_json")}

        raw_rows = flatten_raw_fields(raw_payload)
        if not raw_rows:
            raw_rows = [{"field_path": "", "value": "", "value_json": "null"}]

        for raw_row in raw_rows:
            sheet.append(
                [
                    summary_row.get("task_id"),
                    summary_row.get("file_hash"),
                    raw_row.get("field_path"),
                    raw_row.get("value"),
                    raw_row.get("value_json"),
                    summary_row.get("source_image"),
                    summary_row.get("processed_at"),
                ]
            )

    def _append_raw_json_row(self, workbook, summary_row: dict[str, Any]) -> None:
        if "raw_json" in workbook.sheetnames:
            sheet = workbook["raw_json"]
        else:
            sheet = workbook.create_sheet("raw_json")
            sheet.append(RAW_JSON_HEADERS)

        sheet.append(
            [
                summary_row.get("task_id"),
                summary_row.get("file_hash"),
                summary_row.get("raw_model_json"),
                summary_row.get("source_image"),
                summary_row.get("processed_at"),
            ]
        )

    def _summary_xlsx_target(self) -> Path:
        if not self.summary_xlsx.exists():
            return self.summary_xlsx
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(self.summary_xlsx, read_only=True)
            sheet = workbook.active
            header = [sheet.cell(1, idx + 1).value for idx in range(len(SUMMARY_HEADERS))]
            workbook.close()
        except Exception:
            return self.summary_xlsx.with_name(f"{self.summary_xlsx.stem}_v2{self.summary_xlsx.suffix}")
        if header == SUMMARY_HEADERS:
            return self.summary_xlsx
        return self.summary_xlsx.with_name(f"{self.summary_xlsx.stem}_v2{self.summary_xlsx.suffix}")

    @staticmethod
    def safe_move(source: Path, target_dir: Path, new_name: str | None = None) -> Path:
        target_dir.mkdir(parents=True, exist_ok=True)
        destination = target_dir / (new_name or source.name)
        suffix = 1
        while destination.exists():
            destination = target_dir / f"{source.stem}_{suffix}{source.suffix}"
            suffix += 1
        shutil.move(str(source), str(destination))
        return destination

    @staticmethod
    def dump_json(path: Path, data: dict[str, Any]) -> None:
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
